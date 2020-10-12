# Create a summary report of produce and total sold as shown below:
import pandas as pd
import openpyxl as xl

wb = xl.load_workbook('produceSales.xlsx')

ws = wb.active

maxC = ws.max_column
maxR = ws.max_row

produce_dict = {}

##print("Produce","\t","Cost Per Pound","\t","Amt Sold","\t","Total")
    
for currentrow in ws.iter_rows(min_row = 2, max_row=maxR , max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = round(float(currentrow[2].value),2)
    total = round(float(currentrow[3].value),2)

    produce_dict.setdefault(name,[0.00,0.00,0.00])

    produce_dict[name][0] = round(cost,2)

    
    produce_dict[name][1] += round(amt_sold,2)

    produce_dict[name][2] += round(total,2)

    #print(produce_dict)
    #input()

for k,v in produce_dict.items():
    produce_dict[k] = [v[0],round(v[1]),round(v[2])]

#print(produce_dict)

#!)
df = pd.DataFrame.from_dict(produce_dict)
df.index = ["Cost Per Pound", "Quantity Sold", "Total Sale"]
df1 = df.T.sort_values(by = "Total Sale", ascending = False)
#print(df1)
print("Highest Total Sales:\n", df1.iloc[0])
print()
print("Lowest Total Sales:\n", df1.iloc[39])
print()

#2)
print(df1.loc[["Orange", "Beets"], ["Quantity Sold", "Total Sale"]])
print()
#3) 
print("Total Sale for Cucumbers:", df1.at["Cucumber","Total Sale"])
print()
#4)
df4 = df1.loc[(df1["Quantity Sold"] >= 11500) & (df1["Quantity Sold"] <= 12000)]
print(df4)
print()

#5)
df5 = df4.T
print(df5)





'''  
#output to the screen

for key in produce_dict:
  
  print(key,"\t\t",
       '$',format(produce_dict[key]['cost'],',.2f'),"\t",
        format(produce_dict[key]['amt_sold'],',.2f'),"\t",
        '$',format(produce_dict[key]['total'],',.2f'))   

###write it to a .csv file instead

outfile = open("ProduceReport.csv", 'w')

outfile.write("Produce,Cost Per Pound,Amt Sold,Total\n")

for key in produce_dict:
   outfile.write(key + ',' +
    str(produce_dict[key]['cost']) + ',' +
    format(produce_dict[key]['amt_sold'],',.2f') + ',' +
    format(produce_dict[key]['total'],',.2f') + '\n')

# why would we not want to do the above?
# becos format with ',' will create extra commas and cause problems reading
# the file

for key in produce_dict:
      outfile.write(key + ',' +
        str(produce_dict[key]['cost']) + ',' +
        str(round(produce_dict[key]['amt_sold'],2)) + ',' +
        str(round(produce_dict[key]['total'],2)) + '\n')

#outfile.close()
'''


    

    

    
    
